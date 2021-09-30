#change INPUT_RUN_ID to run the scrolling script on sweetagram again for
#some tag that was already scrolled through, which will pick up new posts and probably 
#some new instagram usernames
INPUT_RUN_ID = "adko_Unique_ID_5" # can be any value, has to be unique
    
#go to sheet & add tags to scrape
INPUT_FILE = "tags_to_scrape1.xlsx"
INPUT_SHEET = "Sheet1"
   

INPUT_DATABASE_NAME = "adkoxMVP_Database.db" # for instagram

#need these from proxy site
INPUT_PROXY_USERNAME = "rnemani" # for instagram
INPUT_PROXY_PASSWORD = "adkoSharedPassword123#" # for instagram

#there is no port for proxies from https://proxyland.io/. leave blank
INPUT_PROXY_PORT = ""

#must be TRUE; Instagram cannot work w/out proxies
INPUT_USE_PROXIES = True # True to use proxies, False to not use them.

#parallel requests
INPUT_BATCH_SIZE = 10 # for sweetagram and instagram


# install: requests, lxml, openpyxl
import requests
from lxml import html
from openpyxl import load_workbook
import os
import sys
import sqlite3
import bz2
import json
import pickle
import time
from datetime import datetime
import threading
import pprint
import random
import re

class Sweetagram_And_Instagram_Scraper:
    def __init__(self, input_run_id, input_filename, input_sheetname, input_databasename, input_proxyusername, input_proxypassword, input_proxyport, input_useproxies, input_batchsize):
        ## check if inputs are good
        self.inputs_are_good = True
        self.is_interrupted = False

        if type(input_run_id) != str:
            print("INPUT_RUN_ID must be a string!")
            self.inputs_are_good = False
        if type(input_filename) != str:
            print("INPUT_FILE must be a string!")
            self.inputs_are_good = False
        if type(input_sheetname) != str:
            print("INPUT_SHEET must be a string!")
            self.inputs_are_good = False
        if type(input_databasename) != str:
            print("INPUT_DATABASE_NAME must be a string!")
            self.inputs_are_good = False
        if type(input_proxyusername) != str:
            print("INPUT_PROXY_USERNAME must be a string!")
            self.inputs_are_good = False
        if type(input_proxypassword) != str:
            print("INPUT_PROXY_PASSWORD must be a string!")
            self.inputs_are_good = False
        if type(input_proxyport) != str:
            print("INPUT_PROXY_PORT must be a string!")
            self.inputs_are_good = False

        if type(input_useproxies) != bool:
            print("INPUT_USE_PROXIES must be True or False")
            self.inputs_are_good = False

        if type(input_batchsize) != int:
            print("INPUT_BATCH_SIZE must be an integer!")
            self.inputs_are_good = False
        else:
            if input_batchsize <= 0:
                print("INPUT_BATCH_SIZE must be a positive integer!")
                self.inputs_are_good = False


        if self.inputs_are_good == False:
            print("Bad inputs, quit!")
            return


        ## if still here, set inputs
        self.run_id = input_run_id
        self.input_file = input_filename
        self.input_sheet = input_sheetname
        self.database_name = input_databasename
        self.proxy_username = input_proxyusername
        self.proxy_password = input_proxypassword
        self.proxy_port = input_proxyport
        self.use_proxies = input_useproxies
        self.batch_size = input_batchsize

        ## create database
        self.db_conn = sqlite3.connect(self.database_name, check_same_thread=False)
        self.db_cursor = self.db_conn.cursor()
        self.db_cursor.execute("CREATE TABLE IF NOT EXISTS SweetagramRuns (tag TEXT NOT NULL, run_id TEXT NOT NULL, items_found INTEGER, PRIMARY KEY(tag, run_id))")
        self.db_cursor.execute("CREATE TABLE IF NOT EXISTS SweetagramPostsTable (post_id TEXT NOT NULL PRIMARY KEY, post_data BLOB, found_under_tag TEXT, found_instagram_username TEXT, username_was_scraped INTEGER)")
        self.db_cursor.execute("CREATE TABLE IF NOT EXISTS InstagramUsersTable (username TEXT NOT NULL PRIMARY KEY, found_under_tag TEXT, json_data BLOB, time_of_scraping TEXT, timestamp REAL)")

        ## create items for threading
        self.good_count = 0
        self.LOCK = threading.Lock()

        ## read inputs
        self.tags_to_scrape = self.read_inputs()
        
        return


    def read_inputs(self):
        try:
            items_to_return = {}
            input_wb = load_workbook(self.input_file)
            input_ws = input_wb[self.input_sheet]
            for row_number in range(2, input_ws.max_row+1):
                potential_tag = input_ws.cell(row=row_number, column=1).value
                if type(potential_tag) == str:
                    if potential_tag not in items_to_return and potential_tag != "":
                        items_to_return[potential_tag] = {"tag":potential_tag, "print_number":len(items_to_return)+1}
                        
            return items_to_return
        except:
            print("An exception while reading inputs - make sure input filename and sheetname are correct!")
            return {}


    def scrape_sweetagram_list_of_usernames_for_input_tags(self):
        if self.inputs_are_good == False or self.is_interrupted == True:
            return

        print("Scraping tags on sweetagram...")
        for tag_to_scrape in self.tags_to_scrape:
            existence_check = self.db_cursor.execute("SELECT EXISTS(SELECT 1 FROM SweetagramRuns WHERE tag=? AND run_id=?)", (self.tags_to_scrape[tag_to_scrape]["tag"], self.run_id)).fetchone()[0]
            if existence_check == 1:
                print("Already scraped tag", self.tags_to_scrape[tag_to_scrape]["tag"], "for run ID", self.run_id)
                continue # already scraped

            # if here, must scrape it!
            # first determine the start url for the script!
            current_url = None # set it in this try/except block
            try:
                start_url_req = requests.get("https://www.sweetagram.com/tag/" + self.tags_to_scrape[tag_to_scrape]["tag"], timeout=60)
                start_tree = html.fromstring(start_url_req.text)
                current_url = "https://www.sweetagram.com/apiget/tag/" + self.tags_to_scrape[tag_to_scrape]["tag"] + "/" +start_tree.xpath("//input[@class='nextpage' and @value]")[0].attrib["value"]
            except KeyboardInterrupt:
                print("Manual interrupt, quit!")
                self.is_interrupted = True
                return
            except:
                print("An exception while trying to find start URL for tag", self.tags_to_scrape[tag_to_scrape]["tag"])
                continue
            
            pagination_scraped = False
            data_to_save = {}
            current_number_of_posts = len(data_to_save)
            page_load_timeout = 120.0
            max_consecutive_empty = 20 # at the end
            current_consecutive_empty = 0
            page_started_scraping_at = time.time()
            
            while 1:
                if time.time() - page_started_scraping_at >= page_load_timeout:
                    break # timeout

                try:
                    r = requests.get(current_url, timeout=25)
                    loaded_json = json.loads(r.text)
                    items_on_this_page = {}
                    if loaded_json["data"] == []:
                        current_consecutive_empty+=1
                        if current_consecutive_empty >= max_consecutive_empty:
                            pagination_scraped = True
                            break
                        page_started_scraping_at = time.time()
                    else:
                        current_consecutive_empty = 0
                        
                    for loaded_post in loaded_json["data"]["edge_hashtag_to_media"]["edges"]:
                        post_id = str(loaded_post["node"]["id"])
                        items_on_this_page[post_id] = ''
                        
                        if post_id not in data_to_save:
                            data_to_save[post_id] = ''

                    print("Currently scraped", len(data_to_save), "posts for tag", self.tags_to_scrape[tag_to_scrape]["tag"], "at", current_url)

                    # insert new items
                    for item_to_insert in items_on_this_page:
                        self.db_cursor.execute("INSERT OR IGNORE INTO SweetagramPostsTable(post_id, found_under_tag, username_was_scraped) VALUES(?,?,?)",
                                               (item_to_insert, self.tags_to_scrape[tag_to_scrape]["tag"], 0))
                    self.db_conn.commit()
                        
                    # set params for next page
                    current_url = current_url[0:current_url.rfind("/")] + "/" + loaded_json["next"]
                    page_started_scraping_at = time.time()
                
                except KeyboardInterrupt:
                    print("Manual interrupt, quit!")
                    self.is_interrupted = True
                    return
                except Exception as e:
                    print("An exception at", current_url, ":", repr(e))
                    continue

            if pagination_scraped == False:
                print("Couldn't scrape pagination for", self.tags_to_scrape[tag_to_scrape]["tag"])
            else: # save data - mark run as complete
                self.db_cursor.execute("INSERT INTO SweetagramRuns(tag, run_id, items_found) VALUES(?,?,?)", (self.tags_to_scrape[tag_to_scrape]["tag"], self.run_id, len(data_to_save) ))
                self.db_conn.commit()
                print("Found a total of", len(data_to_save), "items for tag", self.tags_to_scrape[tag_to_scrape]["tag"])
        return


    def scrape_sweetagram_usernames(self):
        if self.inputs_are_good == False or self.is_interrupted == True:
            return

        print("Scraping usernames from sweetagram...")
        items_to_scrape = self.db_cursor.execute("SELECT post_id, found_under_tag FROM SweetagramPostsTable WHERE username_was_scraped=?", (0,)).fetchall()
        print("Posts left to scrape:", len(items_to_scrape))

        self.good_count = 0
        all_thread_items = []
        for item_to_scrape in items_to_scrape:
            all_thread_items.append({"post_id":item_to_scrape[0], "found_under_tag":item_to_scrape[1]})
            if len(all_thread_items) == self.batch_size:
                ## call it
                all_threads = []
                for a_thread_item in all_thread_items:
                    current_thread = threading.Thread(target=self.sweetagram_post_thread, args=(a_thread_item, ))
                    all_threads.append(current_thread)
                    current_thread.start()

                for thr in all_threads:
                    thr.join()

                print("Current item", items_to_scrape.index(item_to_scrape)+1, "/", len(items_to_scrape), "Good requests in this batch:", self.good_count, "/", len(all_thread_items))
                self.good_count = 0
                all_thread_items = []


        if len(all_thread_items) != 0:
            ## call for residuals
            all_threads = []
            for a_thread_item in all_thread_items:
                current_thread = threading.Thread(target=self.sweetagram_post_thread, args=(a_thread_item, ))
                all_threads.append(current_thread)
                current_thread.start()

            for thr in all_threads:
                thr.join()

            print("Current item", items_to_scrape.index(item_to_scrape)+1, "/", len(items_to_scrape), "Good requests in this batch:", self.good_count, "/", len(all_thread_items))
            self.good_count = 0
            all_thread_items = []
                
        return


    def sweetagram_post_thread(self, input_dict):
        good_to_save = False
        username_to_save = None
        try:
            r = requests.get("https://www.sweetagram.com/posts/viral/#" + input_dict["post_id"], timeout=20)
            pid_to_use = re.findall("(?<=runPosts\(pid,).+?(?=\))", r.text)[0][1:-1]

            # make a second request
            data_req = requests.get("https://www.sweetagram.com/apiget/post/" + input_dict["post_id"] + "?code=" + pid_to_use,
                                    timeout=60)
            loaded_json = json.loads(data_req.text)
            if "error" in loaded_json: # error case
                if loaded_json["error"] == "true":
                    good_to_save = True # don't save any username here!
            else: # normal case
                username_to_save = loaded_json["username"]
                if type(username_to_save) == str:
                    good_to_save = True
        except:
            return

        if good_to_save == True:
            with self.LOCK:
                try:
                    if username_to_save != None:
                        self.db_cursor.execute("INSERT OR IGNORE INTO InstagramUsersTable (username, found_under_tag) VALUES(?,?)",
                                               (username_to_save, input_dict["found_under_tag"]))
                    self.db_cursor.execute("UPDATE SweetagramPostsTable SET post_data=?, found_instagram_username=?, username_was_scraped=? WHERE post_id=?",
                                           (bz2.compress(pickle.dumps(loaded_json)), username_to_save, 1, input_dict["post_id"]))
                    self.db_conn.commit()
                    self.good_count+=1
                except:
                    pass
        return



    def scrape_instagram_data(self):
        if self.inputs_are_good == False or self.is_interrupted == True:
            return

        print("Scraping data from instagram...")
        items_to_scrape = self.db_cursor.execute("SELECT username FROM InstagramUsersTable WHERE json_data IS NULL").fetchall()
        print("Users left to scrape:", len(items_to_scrape))

        self.good_count = 0
        all_thread_items = []
        for item_to_scrape in items_to_scrape:
            all_thread_items.append({"username":item_to_scrape[0]})
            if len(all_thread_items) == self.batch_size:
                ## call it
                all_threads = []
                for a_thread_item in all_thread_items:
                    current_thread = threading.Thread(target=self.instagram_user_thread, args=(a_thread_item, ))
                    all_threads.append(current_thread)
                    current_thread.start()

                for thr in all_threads:
                    thr.join()

                print("Current instagram user", items_to_scrape.index(item_to_scrape)+1, "/", len(items_to_scrape), "Good requests in this batch:", self.good_count, "/", len(all_thread_items))
                self.good_count = 0
                all_thread_items = []


        if len(all_thread_items) != 0:
            ## call for residuals
            all_threads = []
            for a_thread_item in all_thread_items:
                current_thread = threading.Thread(target=self.instagram_user_thread, args=(a_thread_item, ))
                all_threads.append(current_thread)
                current_thread.start()

            for thr in all_threads:
                thr.join()

            print("Current instagram user", items_to_scrape.index(item_to_scrape)+1, "/", len(items_to_scrape), "Good requests in this batch:", self.good_count, "/", len(all_thread_items))
            self.good_count = 0
            all_thread_items = []
            
        return



    def instagram_user_thread(self, input_dict):
        good_to_save = False
        try:
            if self.use_proxies == True:
                proxy_url = self.get_proxyland_proxy_url()
                r = requests.get("https://www.instagram.com/" + input_dict["username"] + "/channel/?__a=1", timeout=10,
                                 headers={"user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36"},
                                 proxies={"http":proxy_url, "https":proxy_url})
            else:
                r = requests.get("https://www.instagram.com/" + input_dict["username"] + "/channel/?__a=1", timeout=10,
                                 headers={"user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36"})
            loaded_data = json.loads( r.text ) # save if good
            probe_data = loaded_data['graphql']['user']['username']
            good_to_save = True # good if here without an exception
        except:
            return

        if good_to_save == True:
            current_time_object = datetime.now()
            with self.LOCK:
                try:
                    self.db_cursor.execute("UPDATE InstagramUsersTable SET json_data=?, time_of_scraping=?, timestamp=? WHERE username=?",
                                           (bz2.compress(pickle.dumps(loaded_data)), current_time_object.strftime("%d-%B-%Y %H:%M:%S"), current_time_object.timestamp(), input_dict["username"]))
                    self.db_conn.commit()
                    self.good_count+=1
                except:
                    pass
        return


    def get_luminati_proxy_url(self, country=None):
        ## country is a <str> if passed. might not be needed for yelp.
        if country == None:
            return 'http://' + self.proxy_username + '-session-' + str(random.random()) + ":" + self.proxy_password + '@zproxy.lum-superproxy.io:'  + str(self.proxy_port)
        else:
            return 'http://' + self.proxy_username + '-country-' + country + '-session-' + str(random.random()) + ":" + self.proxy_password + '@zproxy.lum-superproxy.io:'  + str(self.proxy_port)


    def get_proxyland_proxy_url(self):
        return "http://" + self.proxy_username + ":" + self.proxy_password + "@server.proxyland.io:9090"



if __name__ == '__main__':
    scraper = Sweetagram_And_Instagram_Scraper(INPUT_RUN_ID, INPUT_FILE, INPUT_SHEET, INPUT_DATABASE_NAME, INPUT_PROXY_USERNAME, INPUT_PROXY_PASSWORD, INPUT_PROXY_PORT, INPUT_USE_PROXIES, INPUT_BATCH_SIZE)
    scraper.scrape_sweetagram_list_of_usernames_for_input_tags()
    scraper.scrape_sweetagram_usernames()
    scraper.scrape_instagram_data()
